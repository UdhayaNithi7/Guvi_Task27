from openpyxl import load_workbook

class Orange_data:

    def __init__(self, excel_file, sheet_name):
        self.file = excel_file
        self.sheet = sheet_name
        self.workbook = None

    def load_workbook(self):
        try:
            self.workbook = load_workbook(self.file)
        except Exception as e:
            print(f"Error loading workbook: {e}")

    def row_count(self):
        try:
            if self.workbook is None:
                self.load_workbook()

            if self.workbook is not None:
                sheet = self.workbook[self.sheet]
                return sheet.max_row
            else:
                print("Workbook not loaded.")
                return None
        except Exception as e:
            print(f"Error getting row count: {e}")
            return None

    def column_count(self):
        try:
            if self.workbook is None:
                self.load_workbook()

            if self.workbook is not None:
                sheet = self.workbook[self.sheet]
                return sheet.max_column
            else:
                print("Workbook not loaded.")
                return None
        except Exception as e:
            print(f"Error getting column count: {e}")
            return None

    def access_data(self, row_no, column_no):
        try:
            if self.workbook is None:
                self.load_workbook()

            if self.workbook is not None:
                sheet = self.workbook[self.sheet]
                return sheet.cell(row=row_no, column=column_no).value
            else:
                print("Workbook not loaded.")
                return None
        except Exception as e:
            print(f"Error accessing data: {e}")
            return None
    
    def write_data(self, row_no, column_no, data):
        try:
            if self.workbook is None:
                self.load_workbook()

            if self.workbook is not None:
                sheet = self.workbook[self.sheet]
                sheet.cell(row=row_no, column=column_no).value = data
                self.workbook.save(self.file)
            else:
                print("Workbook not loaded.")
        except Exception as e:
            print(f"Error writing data: {e}")





